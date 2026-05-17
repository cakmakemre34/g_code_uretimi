# -*- coding: utf-8 -*-
"""
STL -> SuperSlicer G-code -> G-code hareketlerinden seçili STL süre tahmini
ve xlsx karşılaştırması.

YENİ ÖZELLİKLER:
- stl_datas/ klasöründeki STL'leri listeler; kullanıcı tek tek seçim yapar.
- Book.xlsx'teki "STL adı -> gerçek süre" eşleşmesini okur.
- Her STL için tahmin edilen süreyi gerçek süreyle karşılaştırır.
- Sonuçları hem terminale basar hem calibration_results.xlsx olarak kaydeder.
- Kalibrasyon sabiti (CORNER_CORRECTION_SEC_PER_WEIGHT ve BASE_JUNCTION_SPEED)
  otomatik olarak hesaplanmaz; elle ayarlanabilir. Ancak kalan hata metrikleri
  (MAE, MAPE, max hata) tabloya eklenir.
"""

import math
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 1) KLASÖR / DOSYA YOLLARI
# ============================================================
BASE = Path(__file__).resolve().parent
STL_DIR = BASE / "stl_datas"
OUT_DIR = BASE / "outputs"
OUT_DIR.mkdir(exist_ok=True)

BASE_INI = BASE / "base_machine.ini"
SUPERSLICER_EXE = Path("/Applications/SuperSlicer.app/Contents/MacOS/SuperSlicer")

# xlsx kalibrasyon verisi (STL adı -> gerçek süre)
REFERENCE_XLSX = BASE / "Book.xlsx"

# ============================================================
# 2) MAKİNE / FIRMWARE SÜRE MODELİ
# ============================================================
BASE_JUNCTION_SPEED_MM_S = 9.886260794376412
CORNER_CORRECTION_THRESHOLD_DEG = 30.0
CORNER_CORRECTION_SEC_PER_WEIGHT = 0.05665627006400997
INCLUDE_E_ONLY_MOVES = True

PRINT_START_MARKERS = [";LAYER:0", ";LAYER_CHANGE"]
PRINT_END_MARKERS = [
    ";Filament-specific end gcode",
    "; filament end gcode",
    "; end gcode",
    ";End of Gcode",
    ";END_PRINT",
]

# ============================================================
# 3) REGEXLER
# ============================================================
INI_KV_RE    = re.compile(r"^\s*([A-Za-z0-9_.-]+)\s*=\s*(.*?)\s*$")
G_MOVE_RE    = re.compile(r"^(G0|G1)\s*(.*)$", re.IGNORECASE)
G92_RE       = re.compile(r"^G92\b(.*)$", re.IGNORECASE)
NUM_RE       = re.compile(r"([XYZEFS])\s*([-+]?\d*\.?\d+(?:[eE][-+]?\d+)?)", re.IGNORECASE)
TYPE_RE      = re.compile(r"^\s*;TYPE:(.+?)\s*$", re.IGNORECASE)
M204_RE      = re.compile(r"^M204\b(.*)$", re.IGNORECASE)
M204_PARAM_RE = re.compile(r"([SPTIA])\s*([-+]?\d*\.?\d+)", re.IGNORECASE)

TYPE_MAP = {
    "WALL INNER": "perimeter",        "INNER PERIMETER": "perimeter",
    "PERIMETER": "perimeter",         "WALL OUTER": "external_perimeter",
    "OUTER PERIMETER": "external_perimeter", "EXTERNAL PERIMETER": "external_perimeter",
    "FILL": "infill",                 "INFILL": "infill",
    "SOLID INFILL": "infill",         "TOP SOLID INFILL": "infill",
    "BRIDGE INFILL": "infill",        "SKIN": "infill",
    "SUPPORT": "infill",              "SUPPORT INTERFACE": "infill",
}

# ============================================================
# 4) YARDIMCILAR
# ============================================================
def die(msg: str, code: int = 1) -> None:
    print(f"\n❌ {msg}")
    sys.exit(code)


def format_duration(seconds: float) -> str:
    total = int(round(seconds))
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    if h:
        return f"{h}s {m:02d}d {s:02d}sn"
    if m:
        return f"{m}d {s:02d}sn"
    return f"{s}sn"


def contains_any(line: str, markers: list[str]) -> bool:
    low = line.lower()
    return any(m.lower() in low for m in markers)


def normalize_type(raw: str) -> str:
    text = raw.strip().upper()
    text = re.sub(r"[\t\-_]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_gcode_words(args: str) -> dict[str, float]:
    return {k.upper(): float(v) for k, v in NUM_RE.findall(args)}


def filament_area_mm2(d_mm: float) -> float:
    return math.pi * (d_mm / 2.0) ** 2


# ============================================================
# 5) XLSX REFERANS VERİSİ OKUMA
# ============================================================
def parse_duration_str(text: str) -> float:
    """
    'X saat Y dk Z sn' / 'X saat Y dk' / 'Y dk Z sn' / 'Xh Ym Zs' vb.
    formatlarını saniyeye çevirir.
    """
    text = str(text).strip().lower()
    # HH:MM:SS
    m = re.fullmatch(r"(\d+):(\d+):(\d+)", text)
    if m:
        return int(m.group(1))*3600 + int(m.group(2))*60 + int(m.group(3))
    # MM:SS
    m = re.fullmatch(r"(\d+):(\d+)", text)
    if m:
        return int(m.group(1))*60 + int(m.group(2))

    total = 0.0
    for val, unit in re.findall(r"([\d.]+)\s*(saat|hour|h|dk|dak|min|m|sn|sek|sec|s)\b", text):
        x = float(val)
        if unit in ("saat", "hour", "h"):
            total += x * 3600
        elif unit in ("dk", "dak", "min", "m"):
            total += x * 60
        elif unit in ("sn", "sek", "sec", "s"):
            total += x
    return total


def load_reference_times(xlsx_path: Path) -> dict[str, float]:
    """
    xlsx'i okur. İlk sütun STL adı (uzantısız veya uzantılı olabilir),
    ikinci sütun gerçek süre metni. {stem -> saniye} döndürür.
    """
    if not xlsx_path.exists():
        print(f"⚠️  Referans xlsx bulunamadı: {xlsx_path}. Karşılaştırma yapılmayacak.")
        return {}

    # header=None: ilk satır da veri olarak okunur
    df = pd.read_excel(xlsx_path, sheet_name=0, header=None)
    refs: dict[str, float] = {}
    for _, row in df.iterrows():
        name_raw = str(row.iloc[0]).strip()
        dur_raw  = str(row.iloc[1]).strip()
        # stem al (uzantıyı kaldır)
        stem = Path(name_raw).stem if "." in name_raw else name_raw
        secs = parse_duration_str(dur_raw)
        if secs > 0:
            refs[stem] = secs
    return refs


# ============================================================
# 6) INI OKUMA/YAZMA
# ============================================================
def parse_ini_keys(text: str) -> set[str]:
    keys: set[str] = set()
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith(("#", ";", "[")):
            continue
        m = INI_KV_RE.match(line)
        if m:
            keys.add(m.group(1))
    return keys


def remove_keys_from_lines(lines: list[str], keys_to_remove: set[str]) -> list[str]:
    if not keys_to_remove:
        return lines[:]
    patterns = [re.compile(rf"^\s*{re.escape(k)}\s*=") for k in keys_to_remove]
    out: list[str] = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith(("#", ";", "[")):
            out.append(line)
            continue
        if not any(p.match(line) for p in patterns):
            out.append(line)
    return out


def override_existing_keys(base_keys, override, candidates, value):
    for k in candidates:
        if k in base_keys:
            override[k] = value


def force_allow_layer_height(base_keys, override, layer_h):
    for k in ["max_layer_height", "max_print_layer_height"]:
        if k in base_keys:
            override[k] = str(layer_h)
    if "min_layer_height" in base_keys:
        override["min_layer_height"] = "0.01"
    if "first_layer_height" in base_keys:
        override["first_layer_height"] = str(layer_h)
    if "layer_height_ranges" in base_keys:
        override["layer_height_ranges"] = f"0.01,{layer_h}"


def build_final_ini(base_ini: Path, override: dict[str, str], out_ini: Path) -> None:
    base_text = base_ini.read_text(encoding="utf-8", errors="ignore")
    lines = base_text.splitlines(keepends=True)
    cleaned = remove_keys_from_lines(lines, set(override.keys()))
    if cleaned and not cleaned[-1].endswith("\n"):
        cleaned.append("\n")
    cleaned.append("\n; ===== PYTHON OVERRIDES =====\n")
    for k, v in override.items():
        cleaned.append(f"{k} = {v}\n")
    out_ini.write_text("".join(cleaned), encoding="utf-8")


# ============================================================
# 7) G-CODE ZAMAN HESABI
# ============================================================
@dataclass
class Segment:
    dist: float
    feed_mm_s: float
    acc_mm_s2: float
    ux: float
    uy: float
    uz: float
    bucket: str
    extruding: bool


@dataclass
class GCodeEstimate:
    raw_kinematic_s: float = 0.0
    corner_correction_s: float = 0.0
    total_s: float = 0.0
    xyze_move_s: float = 0.0
    e_only_s: float = 0.0
    travel_s: float = 0.0
    by_bucket_s: dict = field(default_factory=lambda: defaultdict(float))
    corner_weight: float = 0.0
    sharp_corner_count: int = 0
    total_e_positive_mm: float = 0.0
    filament_volume_mm3: float = 0.0
    filament_weight_g: float = 0.0
    move_count: int = 0
    extrude_move_count: int = 0
    travel_move_count: int = 0
    e_only_count: int = 0
    metadata_time_s: float | None = None


def read_slicer_metadata_time(gcode_path: Path) -> float | None:
    rgx = re.compile(r";\s*estimated printing time.*?=\s*(.+)$", re.IGNORECASE)
    try:
        with gcode_path.open("r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                m = rgx.match(line.strip())
                if m:
                    return parse_duration_str(m.group(1).strip())
    except Exception:
        return None
    return None


def move_time_accel_limited(dist_mm, feed_mm_s, acc_mm_s2,
                             junction_speed_mm_s=BASE_JUNCTION_SPEED_MM_S) -> float:
    d = max(float(dist_mm), 0.0)
    v = max(float(feed_mm_s), 1e-9)
    a = max(float(acc_mm_s2), 1e-9)
    v0 = min(max(float(junction_speed_mm_s), 0.0), v)
    d_acc = max((v*v - v0*v0) / (2.0*a), 0.0)
    if d <= 2.0 * d_acc:
        v_peak = min(math.sqrt(max(v0*v0 + a*d, 0.0)), v)
        if v_peak <= v0 + 1e-12:
            return d / max(v0, 1e-9)
        return 2.0 * (v_peak - v0) / a
    return 2.0 * (v - v0) / a + (d - 2.0*d_acc) / v


def e_only_time(e_mm: float, feed_mm_s: float) -> float:
    return abs(e_mm) / max(feed_mm_s, 1e-9)


def angle_deg_between(a: Segment, b: Segment) -> float:
    dot = max(-1.0, min(1.0, a.ux*b.ux + a.uy*b.uy + a.uz*b.uz))
    return math.degrees(math.acos(dot))


def corner_weight_from_segments(segments: list[Segment]) -> tuple[float, int]:
    threshold = CORNER_CORRECTION_THRESHOLD_DEG
    denom = max(180.0 - threshold, 1e-9)
    weight = 0.0
    sharp_count = 0
    for i in range(1, len(segments)):
        ang = angle_deg_between(segments[i-1], segments[i])
        if ang >= threshold:
            sharp_count += 1
            weight += (ang - threshold) / denom
    return weight, sharp_count


def extract_segments_from_gcode(gcode_path, acc_limits_mm_s2, default_acc_mm_s2,
                                  print_only=True, include_custom=False):
    if not gcode_path.exists():
        die(f"G-code bulunamadı: {gcode_path}")

    section = "start"
    cur_bucket = "extrude_other"
    last_x = last_y = last_z = None
    last_e = None
    last_f_mm_min = None
    relative_e = False

    dyn_acc = {k: float(v) for k, v in acc_limits_mm_s2.items()}
    dyn_default_acc = float(default_acc_mm_s2)

    segments: list[Segment] = []
    est = GCodeEstimate()
    est.metadata_time_s = read_slicer_metadata_time(gcode_path)

    with gcode_path.open("r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue

            if print_only:
                if section == "start" and contains_any(line, PRINT_START_MARKERS):
                    section = "print"
                elif section == "print" and contains_any(line, PRINT_END_MARKERS):
                    section = "end"
            else:
                section = "print"

            if line.startswith("M82"):
                relative_e = False; continue
            if line.startswith("M83"):
                relative_e = True; continue
            if section != "print":
                continue

            g92 = G92_RE.match(line)
            if g92:
                kv = parse_gcode_words(g92.group(1))
                if "X" in kv: last_x = kv["X"]
                if "Y" in kv: last_y = kv["Y"]
                if "Z" in kv: last_z = kv["Z"]
                if "E" in kv: last_e = kv["E"]
                continue

            m204 = M204_RE.match(line)
            if m204:
                params = {k.upper(): float(v) for k, v in M204_PARAM_RE.findall(m204.group(1))}
                if "S" in params:
                    s_val = params["S"]
                    for k in dyn_acc: dyn_acc[k] = s_val
                    dyn_default_acc = s_val
                if "P" in params:
                    p_val = params["P"]
                    for k in ["perimeter", "external_perimeter", "infill", "extrude_other"]:
                        dyn_acc[k] = p_val
                if "T" in params:
                    dyn_acc["travel"] = params["T"]
                continue

            tm = TYPE_RE.match(line)
            if tm:
                t = normalize_type(tm.group(1))
                cur_bucket = ("skip_custom" if t == "CUSTOM" and not include_custom
                              else TYPE_MAP.get(t, "extrude_other"))
                continue

            if cur_bucket == "skip_custom":
                continue

            gm = G_MOVE_RE.match(line)
            if not gm:
                continue

            kv = parse_gcode_words(gm.group(2))
            if "F" in kv:
                last_f_mm_min = kv["F"]
            if last_f_mm_min is None:
                if "X" in kv: last_x = kv["X"]
                if "Y" in kv: last_y = kv["Y"]
                if "Z" in kv: last_z = kv["Z"]
                if "E" in kv:
                    last_e = (last_e or 0) + kv["E"] if relative_e else kv["E"]
                continue

            feed_mm_s = last_f_mm_min / 60.0
            x = kv.get("X", last_x)
            y = kv.get("Y", last_y)
            z = kv.get("Z", last_z)
            dx = 0.0 if (x is None or last_x is None) else x - last_x
            dy = 0.0 if (y is None or last_y is None) else y - last_y
            dz = 0.0 if (z is None or last_z is None) else z - last_z
            dist = math.sqrt(dx*dx + dy*dy + dz*dz)

            de = 0.0
            has_e = "E" in kv
            if has_e:
                if last_e is None:
                    de = kv["E"] if relative_e else 0.0
                    last_e = kv["E"]
                else:
                    if relative_e:
                        de = kv["E"]; last_e += de
                    else:
                        de = kv["E"] - last_e; last_e = kv["E"]

            if "X" in kv: last_x = kv["X"]
            if "Y" in kv: last_y = kv["Y"]
            if "Z" in kv: last_z = kv["Z"]

            if dist <= 1e-9:
                if INCLUDE_E_ONLY_MOVES and has_e and abs(de) > 1e-9:
                    t_e = e_only_time(de, feed_mm_s)
                    est.e_only_s += t_e
                    est.e_only_count += 1
                    if de > 1e-9:
                        est.total_e_positive_mm += de
                continue

            extruding = de > 1e-9
            bucket = cur_bucket if extruding else "travel"
            a_used = float(dyn_acc.get(bucket, dyn_default_acc))
            ux, uy, uz = dx/dist, dy/dist, dz/dist

            segments.append(Segment(dist, feed_mm_s, a_used, ux, uy, uz, bucket, extruding))
            if extruding:
                est.extrude_move_count += 1
                est.total_e_positive_mm += de
            else:
                est.travel_move_count += 1
            est.move_count += 1

    return segments, est


def estimate_gcode_time(gcode_path, acc_limits_mm_s2=None, default_acc_mm_s2=1500.0,
                         filament_d_mm=1.75, filament_density_g_cm3=1.24,
                         print_only=True) -> GCodeEstimate:
    if acc_limits_mm_s2 is None:
        acc_limits_mm_s2 = {
            "perimeter": 1000.0, "external_perimeter": 1000.0,
            "infill": 1500.0, "travel": 1500.0, "extrude_other": 1500.0,
        }

    segments, est = extract_segments_from_gcode(
        gcode_path, acc_limits_mm_s2, default_acc_mm_s2, print_only, not print_only
    )

    for seg in segments:
        t = move_time_accel_limited(seg.dist, seg.feed_mm_s, seg.acc_mm_s2)
        est.xyze_move_s += t
        if seg.bucket == "travel":
            est.travel_s += t
        else:
            est.by_bucket_s[seg.bucket] += t

    cw, sc = corner_weight_from_segments(segments)
    est.corner_weight = cw
    est.sharp_corner_count = sc
    est.corner_correction_s = cw * CORNER_CORRECTION_SEC_PER_WEIGHT
    est.raw_kinematic_s = est.xyze_move_s + est.e_only_s
    est.total_s = est.raw_kinematic_s + est.corner_correction_s

    vol_mm3 = est.total_e_positive_mm * filament_area_mm2(filament_d_mm)
    est.filament_volume_mm3 = vol_mm3
    est.filament_weight_g = vol_mm3 * filament_density_g_cm3 / 1000.0
    return est


# ============================================================
# 8) SUPERSLICER ÇALIŞTIRMA
# ============================================================
def run_slicer(final_ini_path, stl_path, gcode_path) -> bool:
    cmd = [
        str(SUPERSLICER_EXE), "--load", str(final_ini_path),
        "--export-gcode", "--output", str(gcode_path), str(stl_path),
    ]
    res = subprocess.run(
        cmd, cwd=str(SUPERSLICER_EXE.parent),
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
    )
    if res.returncode != 0:
        print(f"   ⚠️ SuperSlicer stderr: {res.stderr[:300]}")
    return gcode_path.exists() and gcode_path.stat().st_size > 0


# ============================================================
# 9) SLICER AYARLARI (interaktif, bir kez sorulur)
# ============================================================
def ask(prompt, default, cast=float, lo=None, hi=None):
    while True:
        raw = input(f"  {prompt} [{default}]: ").strip().replace(",", ".")
        if not raw:
            val = cast(default)
        else:
            try:
                val = cast(raw)
            except Exception:
                print("  Geçersiz giriş, tekrar dene.")
                continue
        if lo is not None and val < lo:
            print(f"  ⚠️  {val} önerilen min ({lo}) altında.")
        if hi is not None and val > hi:
            print(f"  ⚠️  {val} önerilen max ({hi}) üstünde.")
        return val


def ask_slicer_settings():
    print("\n════════════════════════════════════════════")
    print("  Dilimleyici Ayarları (tüm STL'ler için ortak)")
    print("════════════════════════════════════════════")
    print("  (Ender-3 + PLA varsayılanları)")

    nozzle_temp = ask("Nozzle Sıcaklığı (°C)", 210, int, 190, 230)
    bed_temp    = ask("Tabla Sıcaklığı (°C)",  60,  int, 45,  70)
    nozzle_d    = ask("Nozzle Çapı (mm)",       0.4, float, 0.2, 1.2)
    layer_h     = ask("Katman Yüksekliği (mm)", 0.2, float, 0.08, round(0.75*nozzle_d, 3))
    perimeters  = ask("Duvar Sayısı",           2,   int, 1, 8)
    infill_pct  = ask("Infill (%)",             20,  int, 0, 100)
    perim_sp    = ask("Perimeter Hızı (mm/s)",  60,  float, 20, 120)
    ext_sp      = ask("External Perim. Hızı",   40,  float, 15, 100)
    infill_sp   = ask("Infill Hızı (mm/s)",     80,  float, 20, 500)
    travel_sp   = ask("Travel Hızı (mm/s)",     180, float, 80, 600)
    def_acc     = ask("Default Accel (mm/s²)",  1500, int, 300, 5000)
    per_acc     = ask("Perimeter Accel",        1000, int, 300, 4000)
    inf_acc     = ask("Infill Accel",           1500, int, 300, 6000)
    vol_limit   = ask("Max Volumetric (mm³/s)", 12.0, float, 4.0, 25.0)

    acc_limits = {
        "perimeter": float(per_acc), "external_perimeter": float(per_acc),
        "infill": float(inf_acc),    "travel": float(def_acc),
        "extrude_other": float(def_acc),
    }
    speed_limits = {
        "perimeter": float(perim_sp), "external_perimeter": float(ext_sp),
        "infill": float(infill_sp),   "travel": float(travel_sp),
    }

    settings = dict(
        nozzle_temp=nozzle_temp, bed_temp=bed_temp, nozzle_d=nozzle_d,
        layer_h=layer_h, perimeters=perimeters, infill_pct=infill_pct,
        perim_sp=perim_sp, ext_sp=ext_sp, infill_sp=infill_sp,
        travel_sp=travel_sp, def_acc=def_acc, per_acc=per_acc,
        inf_acc=inf_acc, vol_limit=vol_limit,
    )
    return settings, acc_limits, speed_limits


def build_ini_override(base_keys, settings):
    s = settings
    override: dict[str, str] = {}
    override_existing_keys(base_keys, override,
        ["temperature", "first_layer_temperature", "nozzle_temperature",
         "print_temperature", "print_first_layer_temperature"], str(s["nozzle_temp"]))
    override_existing_keys(base_keys, override,
        ["bed_temperature", "first_layer_bed_temperature", "bed_temperature_0",
         "first_layer_bed_temperature_0"], str(s["bed_temp"]))

    if "nozzle_diameter" in base_keys:
        override["nozzle_diameter"] = str(s["nozzle_d"])

    auto_width = round(1.10 * s["nozzle_d"], 3)
    for k in ["extrusion_width", "perimeter_extrusion_width",
              "external_perimeter_extrusion_width", "infill_extrusion_width",
              "solid_infill_extrusion_width", "top_infill_extrusion_width"]:
        if k in base_keys:
            override[k] = str(auto_width)

    if "layer_height"  in base_keys: override["layer_height"]  = str(s["layer_h"])
    if "perimeters"    in base_keys: override["perimeters"]    = str(s["perimeters"])

    base_text = BASE_INI.read_text(encoding="utf-8", errors="ignore")
    for k in ["infill_density", "fill_density"]:
        if k in base_keys:
            line = next((ln for ln in base_text.splitlines() if ln.strip().startswith(k)), "")
            override[k] = f"{s['infill_pct']}%" if "%" in line else str(s["infill_pct"] / 100.0)

    for k, v in [("perimeter_speed", s["perim_sp"]), ("external_perimeter_speed", s["ext_sp"]),
                 ("infill_speed", s["infill_sp"]), ("solid_infill_speed", s["infill_sp"]),
                 ("top_solid_infill_speed", s["infill_sp"]), ("travel_speed", s["travel_sp"])]:
        if k in base_keys:
            override[k] = str(v)

    for k, v in [("default_acceleration", s["def_acc"]), ("perimeter_acceleration", s["per_acc"]),
                 ("external_perimeter_acceleration", s["per_acc"]),
                 ("infill_acceleration", s["inf_acc"]), ("solid_infill_acceleration", s["inf_acc"]),
                 ("top_solid_infill_acceleration", s["inf_acc"]),
                 ("travel_acceleration", s["def_acc"])]:
        if k in base_keys:
            override[k] = str(v)

    for k in ["filament_max_volumetric_speed", "max_volumetric_speed"]:
        if k in base_keys:
            override[k] = str(s["vol_limit"])

    force_allow_layer_height(base_keys, override, s["layer_h"])
    return override


# ============================================================
# 10) EXCEL RAPORU
# ============================================================
def seconds_to_hms_str(s: float) -> str:
    return format_duration(s)


def write_results_xlsx(rows: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sonuçlar"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ok_fill     = PatternFill("solid", start_color="C6EFCE")
    warn_fill   = PatternFill("solid", start_color="FFEB9C")
    bad_fill    = PatternFill("solid", start_color="FFC7CE")
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left", vertical="center")
    thin        = Side(style="thin", color="AAAAAA")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "STL Dosyası", "Tahmin Edilen Süre", "Gerçek Süre",
        "Fark (sn)", "Fark (%)", "Durum",
        "Ham Kinematik (sn)", "Köşe Düzeltme (sn)", "Köşe Ağırlığı",
        "Keskin Köşe Sayısı", "Filament (g)",
    ]
    col_widths = [42, 22, 20, 14, 12, 12, 22, 22, 18, 20, 15]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(rows, start=2):
        pct = row.get("pct_err")
        if pct is None:
            fill = PatternFill()
        elif abs(pct) <= 2:
            fill = ok_fill
        elif abs(pct) <= 5:
            fill = warn_fill
        else:
            fill = bad_fill

        values = [
            row["stem"],
            row["est_str"],
            row.get("ref_str", "—"),
            round(row["diff_s"], 1) if row.get("diff_s") is not None else "—",
            f"{pct:+.2f}%" if pct is not None else "—",
            row.get("status", "—"),
            round(row["raw_s"], 1),
            round(row["corner_s"], 1),
            round(row["corner_w"], 3),
            row["sharp_cnt"],
            round(row["filament_g"], 2),
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = left if ci == 1 else center
            if ci >= 2:
                cell.fill = fill
        ws.row_dimensions[ri].height = 18

    # Özet istatistikler
    valid = [r for r in rows if r.get("pct_err") is not None]
    if valid:
        summary_row = len(rows) + 3
        ws.cell(row=summary_row, column=1, value="📊 Özet İstatistikler").font = Font(bold=True, name="Arial")
        mae  = sum(abs(r["diff_s"]) for r in valid) / len(valid)
        mape = sum(abs(r["pct_err"]) for r in valid) / len(valid)
        max_err = max(abs(r["pct_err"]) for r in valid)
        for label, val in [("MAE (sn)", f"{mae:.1f}"), ("MAPE (%)", f"{mape:.2f}"),
                            ("Maks. Hata (%)", f"{max_err:.2f}")]:
            summary_row += 1
            ws.cell(row=summary_row, column=1, value=label).font = Font(name="Arial")
            ws.cell(row=summary_row, column=2, value=val).font = Font(name="Arial")

    wb.save(out_path)
    print(f"\n💾 Sonuçlar kaydedildi: {out_path}")


# ============================================================
# 11) TEKLİ İŞLEM: STL listesinden seçim yap
# ============================================================
def choose_stl_from_list(stls: list[Path]) -> Path:
    """stl_datas klasöründeki STL dosyalarını numaralı listeler ve kullanıcıya seçtirir."""
    print("\n📂 STL Listesi")
    print("════════════════════════════════════════════")
    for i, stl in enumerate(stls, start=1):
        print(f"  [{i}] {stl.name}")

    while True:
        raw = input("\nSeçilecek STL numarası: ").strip()
        try:
            idx = int(raw)
            if 1 <= idx <= len(stls):
                return stls[idx - 1]
        except Exception:
            pass
        print(f"  Geçersiz seçim. 1 ile {len(stls)} arasında numara gir.")


def process_selected_stl(settings, acc_limits, speed_limits) -> None:
    if not BASE_INI.exists():
        die(f"base_machine.ini bulunamadı: {BASE_INI}")
    if not SUPERSLICER_EXE.exists():
        die(f"SuperSlicer bulunamadı: {SUPERSLICER_EXE}")
    if not STL_DIR.exists():
        die(f"'{STL_DIR}' klasörü yok.")

    stls = sorted(p for p in STL_DIR.iterdir() if p.suffix.lower() == ".stl")
    if not stls:
        die("stl_datas/ içinde STL dosyası bulunamadı.")

    selected_stl = choose_stl_from_list(stls)
    print(f"\n✅ Seçilen STL: {selected_stl.name}")

    ref_times = load_reference_times(REFERENCE_XLSX)
    print(f"\n📋 Yüklenen referans süreleri: {len(ref_times)} adet")
    if selected_stl.stem in ref_times:
        print(f"   Bu STL için ölçülen süre bulundu: {format_duration(ref_times[selected_stl.stem])}")
    else:
        print("   Bu STL için Book.xlsx içinde ölçülen süre bulunamadı.")

    base_text = BASE_INI.read_text(encoding="utf-8", errors="ignore")
    base_keys = parse_ini_keys(base_text)
    override = build_ini_override(base_keys, settings)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_ini_path = OUT_DIR / f"config_{stamp}.ini"
    build_final_ini(BASE_INI, override, final_ini_path)
    print(f"\n⚙️  Ortak config oluşturuldu: {final_ini_path.name}")

    gcode_path = OUT_DIR / f"{selected_stl.stem}_{stamp}.gcode"

    print("\n🔄 Seçili STL işleniyor...\n")
    print(f"STL: {selected_stl.name}")

    ok = run_slicer(final_ini_path, selected_stl, gcode_path)
    if not ok:
        die("Slice başarısız. G-code üretilemedi.")

    est = estimate_gcode_time(
        gcode_path=gcode_path,
        acc_limits_mm_s2=acc_limits,
        default_acc_mm_s2=float(acc_limits.get("travel", 1500.0)),
        filament_d_mm=1.75,
        print_only=True,
    )

    ref_s = ref_times.get(selected_stl.stem)
    diff_s = (est.total_s - ref_s) if ref_s else None
    pct = (diff_s / ref_s * 100.0) if (ref_s and ref_s > 0) else None

    if pct is None:
        status = "Referans yok"
    elif abs(pct) <= 2:
        status = "✅ İyi"
    elif abs(pct) <= 5:
        status = "🟡 Kabul"
    else:
        status = "❌ Yüksek hata"

    est_str = format_duration(est.total_s)
    ref_str = format_duration(ref_s) if ref_s else "—"

    print("\n========== SONUÇ ==========")
    print(f"STL              : {selected_stl.name}")
    print(f"G-code           : {gcode_path.name}")
    print(f"Tahmini süre     : {est_str} ({est.total_s:.2f} sn)")
    print(f"Ham kinematik    : {format_duration(est.raw_kinematic_s)} ({est.raw_kinematic_s:.2f} sn)")
    print(f"Köşe düzeltmesi  : {est.corner_correction_s:.2f} sn")
    print(f"Filament tahmini : {est.filament_weight_g:.2f} g")

    if ref_s:
        sign = "+" if diff_s >= 0 else ""
        print(f"Ölçülen süre     : {ref_str} ({ref_s:.2f} sn)")
        print(f"Fark             : {sign}{diff_s:.0f} sn ({sign}{pct:.2f}%) {status}")
    else:
        print("Ölçülen süre     : Book.xlsx içinde bu STL için bulunamadı")

    row = dict(
        stem=selected_stl.stem,
        est_str=est_str,
        est_s=est.total_s,
        ref_str=ref_str,
        ref_s=ref_s,
        diff_s=diff_s,
        pct_err=pct,
        status=status,
        raw_s=est.raw_kinematic_s,
        corner_s=est.corner_correction_s,
        corner_w=est.corner_weight,
        sharp_cnt=est.sharp_corner_count,
        filament_g=est.filament_weight_g,
    )

    out_xlsx = OUT_DIR / f"single_result_{selected_stl.stem}_{stamp}.xlsx"
    write_results_xlsx([row], out_xlsx)


# ============================================================
# 12) MAIN
# ============================================================
def main() -> None:
    print("\n╔══════════════════════════════════════════════════╗")
    print("║  STL Seçimli Süre Tahmini + xlsx Karşılaştırma   ║")
    print("╚══════════════════════════════════════════════════╝")

    settings, acc_limits, speed_limits = ask_slicer_settings()
    process_selected_stl(settings, acc_limits, speed_limits)


if __name__ == "__main__":
    main()